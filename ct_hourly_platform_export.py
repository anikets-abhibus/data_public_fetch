import os
import sys
import time
import json
import datetime as dt
import requests
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook, Workbook
import pyzipper

# --------------- CONFIG ---------------
CT_ACCOUNT_ID = os.environ.get("CT_ACCOUNT_ID")
CT_PASSCODE = os.environ.get("CT_PASSCODE")
SLACK_WEBHOOK = os.environ.get("SLACK_WEBHOOK")
ZIP_PASSWORD = os.environ.get("ZIP_PASSWORD")

URL = "https://eu1.api.clevertap.com/1/counts/profiles.json"
# CHANGED FOR PREVIOUS 6 MONTHS: (April 19, 2025 to October 15, 2025)
DATE_FROM = dt.date(2025, 4, 19)
DATE_TO = dt.date(2025, 10, 15)
OUTPUT_FILE = "ct_hourly_platform_previous_6months.xlsx"
ZIP_FILE = "ct_hourly_platform_previous_6months.zip"
LOCAL_PATH = f"/tmp/{OUTPUT_FILE}"
ZIP_PATH = f"/tmp/{ZIP_FILE}"

# GitHub Actions max runtime is 6 hours. We stop safely at 5.5 hours.
MAX_RUNTIME_SECONDS = 5.5 * 3600
START_TIME = time.time()
# --------------------------------------

missing = []
if not CT_ACCOUNT_ID: missing.append("CT_ACCOUNT_ID")
if not CT_PASSCODE: missing.append("CT_PASSCODE")
if not SLACK_WEBHOOK: missing.append("SLACK_WEBHOOK")
if not ZIP_PASSWORD: missing.append("ZIP_PASSWORD")

if missing:
    print(f"Missing required environment variables: {', '.join(missing)}. Exiting.")
    sys.exit(1)

HEADERS = {
    "X-CleverTap-Account-Id": CT_ACCOUNT_ID,
    "X-CleverTap-Passcode": CT_PASSCODE,
    "Content-Type": "application/json",
}

EVENTS = [
    "bus_home", "srp_landing", "bus_selected_onward", "seat_selection_onward",
    "bus_boarding_onward", "bus_dropping_onward", "booking_details",
    "Charged", "cancellations", "no_results",
]

PLATFORMS = [
    ("platform", "", "0.Grand Total"),
    ("platform", "android app", "1.Android"),
    ("platform", "PWA", "9.PWA"),
    ("platform", "ios app", "2.iOS"),
    ("platform", "ab-mobile-new", "3.New Mobile website"),
    ("platform", "ab-website-new", "4.New Desktop website"),
    ("platform", "ixiweb", "ixigo New Desktop"),
    ("platform_name", "ixigomsite", "8.ixigo mobile website"),
    ("platform_name", "Gpay", "Gpay"),
    ("platform_name", "ixflights", "7.ixigo flights"),
    ("platform_name", "confirmtkt", "6.Confirmtkt"),
    ("platform_name", "ixtrains", "5.ixigo trains"),
    ("platform_name", "phonepe", "phonepe"),
    ("platform_name", "ubermweb", "ubermweb"),
]

BATCH_SIZE = 120
POLL_WAIT = 15
RETRY_WAIT = 15
MAX_HTTP_RETRIES = 3
DAY_COOLDOWN = 3
WORKERS = 2
BACKOFF_BASE = 20

COLUMNS = ["date", "hour", "platform"] + [f"{ev}_users" for ev in EVENTS]


def log(msg):
    ts = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def slack_notify(text):
    try:
        requests.post(SLACK_WEBHOOK, json={"text": text}, timeout=10)
    except Exception as e:
        log(f"Slack notify failed: {e}")


def extract_zip():
    if os.path.exists(ZIP_PATH):
        log(f"Found existing encrypted zip {ZIP_PATH}. Extracting...")
        try:
            with pyzipper.AESZipFile(ZIP_PATH) as zf:
                zf.pwd = ZIP_PASSWORD.encode('utf-8')
                zf.extractall(path="/tmp/")
            log("Extraction successful.")
        except Exception as e:
            log(f"Failed to extract zip (password wrong or corrupt): {e}")
            sys.exit(1)
    else:
        log("No existing zip found. Starting fresh.")


def create_encrypted_zip():
    log(f"Creating encrypted zip {ZIP_PATH}...")
    if os.path.exists(LOCAL_PATH):
        with pyzipper.AESZipFile(ZIP_PATH, 'w', compression=pyzipper.ZIP_LZMA, encryption=pyzipper.WZ_AES) as zf:
            zf.pwd = ZIP_PASSWORD.encode('utf-8')
            zf.write(LOCAL_PATH, arcname=OUTPUT_FILE)
        log("Encrypted zip created successfully.")
    else:
        log("No Excel file found to zip.")


def post_with_retry(url, json_body=None, data=None):
    for attempt in range(MAX_HTTP_RETRIES):
        try:
            kw = {"headers": HEADERS, "timeout": 60}
            if json_body is not None:
                kw["json"] = json_body
            else:
                kw["data"] = data or ""
            r = requests.post(url, **kw)
            if r.status_code == 429:
                wait = BACKOFF_BASE * (2 ** attempt)
                log(f"    429 backoff {wait}s (attempt {attempt+1})")
                time.sleep(wait)
                continue
            return r.json()
        except Exception as e:
            wait = BACKOFF_BASE * (2 ** attempt)
            if attempt < MAX_HTTP_RETRIES - 1:
                log(f"    err: {e} retry in {wait}s")
                time.sleep(wait)
            else:
                log(f"    err: {e} giving up")
    return {"status": "fail", "error": "max retries"}


def load_completed_dates(path):
    done = set()
    if not os.path.exists(path):
        return done
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        date_counts = {}
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            d = row[0]
            if d:
                ds = str(d)
                date_counts[ds] = date_counts.get(ds, 0) + 1
        wb.close()
        expected = 24 * len(PLATFORMS)
        for d, count in date_counts.items():
            if count >= expected:
                done.add(d)
    except Exception as e:
        log(f"Warning reading existing Excel: {e}")
    return done


def append_to_excel(path, new_rows):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "CT_Export"
        ws.append(COLUMNS)
    for row in new_rows:
        ws.append([row.get(c, "") for c in COLUMNS])
    wb.save(path)


def build_day_queries(date_int):
    queries = []
    for event in EVENTS:
        for h in range(24):
            sh = f"{h:02d}:00"
            eh = f"{h:02d}:59"
            for prop_name, prop_value, display in PLATFORMS:
                payload = {
                    "event_name": event,
                    "event_properties": [
                        {"name": prop_name, "operator": "equals", "value": prop_value}
                    ],
                    "session_properties": [
                        {"name": "time_of_day", "value": [sh, eh]}
                    ],
                    "from": date_int,
                    "to": date_int,
                }
                queries.append({
                    "payload": payload,
                    "event": event,
                    "hour": sh,
                    "platform": display,
                })
    return queries


def fire_sub_batch(queries):
    tokens = []
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = [(q, ex.submit(post_with_retry, URL, json_body=q["payload"])) for q in queries]
        for q, f in futures:
            try:
                j = f.result(timeout=120)
            except Exception:
                j = {}
            tokens.append({
                "event": q["event"], "hour": q["hour"],
                "platform": q["platform"], "req_id": j.get("req_id"),
            })
    return tokens


def poll_tokens(tokens):
    results = {}
    to_poll = []
    for t in tokens:
        if not t["req_id"]:
            results[(t["event"], t["hour"], t["platform"])] = 0
        else:
            to_poll.append(t)

    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futures = [(t, ex.submit(post_with_retry, f"{URL}?req_id={t['req_id']}", data="")) for t in to_poll]
        for t, f in futures:
            try:
                j = f.result(timeout=120)
            except Exception:
                j = {}
            if j.get("status") == "partial":
                results[(t["event"], t["hour"], t["platform"])] = None
            else:
                results[(t["event"], t["hour"], t["platform"])] = j.get("count", 0)

    still_partial = [t for t in to_poll if results.get((t["event"], t["hour"], t["platform"])) is None]
    final_results = {k: (v if v is not None else 0) for k, v in results.items()}
    return final_results, still_partial


def process_day(date_int, day_str):
    queries = build_day_queries(date_int)
    total_batches = (len(queries) + BATCH_SIZE - 1) // BATCH_SIZE
    log(f"  {len(queries)} queries, {total_batches} batches")

    all_results = {}

    for i in range(0, len(queries), BATCH_SIZE):
        batch = queries[i:i + BATCH_SIZE]
        bn = i // BATCH_SIZE + 1
        log(f"  batch {bn}/{total_batches} ({len(batch)}q) fire...")

        tokens = fire_sub_batch(batch)

        log(f"  batch {bn} wait {POLL_WAIT}s...")
        time.sleep(POLL_WAIT)

        log(f"  batch {bn} poll...")
        results, partial = poll_tokens(tokens)
        all_results.update(results)

        if partial:
            log(f"  batch {bn}: {len(partial)} partial, retry in {RETRY_WAIT}s")
            time.sleep(RETRY_WAIT)
            r2, p2 = poll_tokens(partial)
            all_results.update(r2)
            if p2:
                log(f"  batch {bn}: {len(p2)} still partial, retry2...")
                time.sleep(RETRY_WAIT)
                r3, p3 = poll_tokens(p2)
                all_results.update(r3)
                for t in p3:
                    all_results[(t["event"], t["hour"], t["platform"])] = 0

    rows = []
    for h in range(24):
        hh = f"{h:02d}:00"
        for _, _, display in PLATFORMS:
            row = {"date": day_str, "hour": hh, "platform": display}
            for ev in EVENTS:
                row[f"{ev}_users"] = all_results.get((ev, hh, display), 0)
            rows.append(row)
    return rows


def main():
    log("Initializing...")
    
    # 1. Extract existing zip if it exists (downloaded by GitHub Actions step)
    extract_zip()

    completed = load_completed_dates(LOCAL_PATH)

    all_dates = []
    d = DATE_FROM
    while d <= DATE_TO:
        all_dates.append(d)
        d += dt.timedelta(days=1)

    remaining = [d for d in all_dates if d.isoformat() not in completed]
    total = len(all_dates)
    skip = total - len(remaining)

    if not remaining:
        log("All dates already processed. Exiting.")
        slack_notify("*CT Export COMPLETE!*\nAll 180 days exported.")
        create_encrypted_zip() # Ensure zip is ready for final upload
        return

    log("=" * 60)
    log("CleverTap Hourly Platform Export (GitHub Artifacts - Encrypted)")
    log(f"Range: {DATE_FROM} to {DATE_TO} ({total} days)")
    log(f"Done: {skip} | Remaining: {len(remaining)}")
    log("=" * 60)

    start_msg = (
        f"*CT Export Run Started*\n"
        f"Already done: {skip}/{total} days\n"
        f"Will run for up to 5.5 hours."
    )
    slack_notify(start_msg)

    times = []
    days_processed_this_run = 0

    try:
        for idx, day in enumerate(remaining):
            # Check runtime limit
            if time.time() - START_TIME > MAX_RUNTIME_SECONDS:
                log("Reached 5.5 hour safe shutdown limit.")
                break

            day_str = day.isoformat()
            date_int = int(day.strftime("%Y%m%d"))
            t0 = time.time()

            log(f"\n--- Day {skip + idx + 1}/{total}: {day_str} ---")

            rows = process_day(date_int, day_str)
            append_to_excel(LOCAL_PATH, rows)
            
            # Create encrypted zip after each day so if it crashes, the latest zip is ready for the upload step
            create_encrypted_zip()

            days_processed_this_run += 1
            elapsed = time.time() - t0
            times.append(elapsed)
            avg = sum(times) / len(times)
            left = len(remaining) - idx - 1

            log(f"  {elapsed:.0f}s | avg {avg:.0f}s/day | {left} left")

            if left > 0:
                time.sleep(DAY_COOLDOWN)

        # Final status
        total_done = skip + days_processed_this_run
        if total_done >= total:
            done_msg = (
                f"*CT Export COMPLETE!*\n"
                f"All {total} days exported successfully."
            )
            log("\n=== EXPORT COMPLETE ===")
            slack_notify(done_msg)
        else:
            pause_msg = (
                f"*CT Export Paused (Time Limit)*\n"
                f"Processed {days_processed_this_run} days this run.\n"
                f"Total progress: {total_done}/{total} days.\n"
                f"Will resume automatically on next schedule."
            )
            log("\n=== EXPORT PAUSED ===")
            slack_notify(pause_msg)

    except Exception as e:
        err_msg = (
            f"*CT Export ERROR*\n"
            f"Error: {e}\n"
            f"Progress saved up to last completed day."
        )
        log(f"ERROR: {e}")
        slack_notify(err_msg)
        create_encrypted_zip() # Try to save what we have
        raise

if __name__ == "__main__":
    main()
